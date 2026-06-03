import openpyxl
import re
import json
import datetime
from pathlib import Path

EXCEL_FILE = "data/опт 01.06..xlsx"
JSON_FILE = "data/products.json"

if not Path(EXCEL_FILE).exists():
    print(f"❌ Файл не найден: {EXCEL_FILE}")
    exit(1)

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
sheet = wb.active

START_ROW = 8
NAME_COL = 5      # E
PRICE_COL = 14    # N
STOCK_COL = 15    # O

products = []
pid = 1

for row_idx in range(START_ROW, sheet.max_row + 1):
    name = sheet.cell(row=row_idx, column=NAME_COL).value
    price_cell = sheet.cell(row=row_idx, column=PRICE_COL).value
    stock_cell = sheet.cell(row=row_idx, column=STOCK_COL).value

    if not name or not isinstance(name, str):
        continue
    name = name.strip()
    if len(name) < 3:
        continue
    if any(x in name for x in ['Итого', 'Всего', 'Прайс-лист']):
        continue

    try:
        if price_cell is None or stock_cell is None:
            continue
        if isinstance(price_cell, (int, float)):
            price_val = float(price_cell)
        else:
            price_val = float(str(price_cell).replace(',', '.').replace(' ', ''))
        if isinstance(stock_cell, (int, float)):
            stock_val = float(stock_cell)
        else:
            stock_val = float(str(stock_cell).replace(',', '').replace(' ', ''))
    except:
        continue

    # Категория
    name_low = name.lower()
    if 'банк' in name_low and 'бутыл' not in name_low:
        category = 'Стеклянные банки'
    elif 'бутылк' in name_low or 'бутыль' in name_low:
        category = 'Бутылки и бутыли'
    elif 'крышк' in name_low or 'колпач' in name_low:
        category = 'Крышки и колпачки'
    else:
        category = 'Разное'

    vol_match = re.search(r'(\d+[.,]?\d*)\s*(л|мл|литр)', name, re.IGNORECASE)
    volume = f"{vol_match.group(1).replace(',', '.')} {vol_match.group(2).lower()}" if vol_match else ''
    diam_match = re.search(r'd(\d+)', name)
    diameter = f"d{diam_match.group(1)}" if diam_match else ''

    # НОВОЕ ОПИСАНИЕ (без доставки, с условием опта)
    description = f"✨ {name} ✨\n\n💰 Цена: {price_val:.2f} руб. (опт, с НДС)\n📦 В наличии: {stock_val:.0f} шт.\n"
    if volume:
        description += f"📏 Объём: {volume}\n"
    if diameter:
        description += f"🔘 Диаметр горла: {diameter}\n"
    description += "💰 Оптовая цена на стеклотару действует при заказе от 5000 ₽\n⭐ Идеально для консервации, виноделия и домашнего хозяйства."

    images = [f"https://via.placeholder.com/400x300?text={name[:30]}" for _ in range(3)]

    products.append({
        "id": pid,
        "name": name,
        "price": price_val,
        "stock": stock_val,
        "volume": volume,
        "diameter": diameter,
        "category": category,
        "description": description,
        "images": images
    })
    pid += 1

# Мета-информация с датой
output_data = {
    "last_updated": datetime.datetime.now().strftime("%d.%m.%Y"),
    "total_items": len(products),
    "items": products
}

Path("data").mkdir(exist_ok=True)
with open(JSON_FILE, 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"✅ Обработано товаров: {len(products)}")
print(f"📅 Дата обновления: {output_data['last_updated']}")